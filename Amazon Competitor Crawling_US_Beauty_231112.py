from selenium import webdriver
from bs4 import BeautifulSoup
import requests
import random
# import schedule
import time
from datetime import datetime
import pandas as pd
import xlsxwriter
import os
from openpyxl import load_workbook
from selenium import webdriver
from selenium.webdriver.chrome.service import Service as ChromeService
from selenium.webdriver.chrome.options import Options
from webdriver_manager.chrome import ChromeDriverManager

# def playing():

## Chrome driver / header setting ##
service = ChromeService(excecutable_path="C:/Users\juntaehwang\Desktop\새 폴더\Programs\chromedriver_win32\chromedriver.exe")
options = webdriver.ChromeOptions()
driver = webdriver.Chrome(service=service, options=options)
# driver = webdriver.Chrome(excecutable_path="C:/Users\juntaehwang\Desktop\새 폴더\Programs\chromedriver_win32\chromedriver.exe")

def Request_page(page):
    header = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/119.0.0.0 Safari/537.36',
        'referer': page}
    cookie = {}
    response = requests.get(page, cookies=cookie, headers=header)
    soup = BeautifulSoup(response.content, "html.parser")
    time.sleep(random.randint(1, 10))

    if response.status_code == 200:
        print()
        return response
        print("crawling is just started :)")
    else:
        return "Error"

# Save data into excel file
# making a data frame with a dictionary form
# def save_excel(sp_button, brand, asin_list, title_list, price, ratings, rank):
def save_excel(asin_list, title_list):
    rev={
        # 'Date': current_date,
        # 'SP_Button': sp_button,
        # 'Brand': brand,
        'ASIN': asin_list,
        'Title': title_list,
        # 'Price': price,
        # 'Rating': ratings,
        # 'Rank': rank
        }

    # matching the count of columns and rows
    df_1 = pd.DataFrame.from_dict(rev, orient='index')
    df_1 = df_1.transpose()

    df_1.head(5)
    df_1.shape
    # df_1.sort_values(by='Date')
    df_1.set_index = current_date

    print(df_1)

    ## open existing excel file without overwriting ## - 3 (writer line inside try block)
    # https://stackoverflow.com/questions/66531396/export-pandas-dataframe-to-xlsx-dealing-with-the-openpyxl-issue-on-python-3-9
    def append_df_to_excel(filename, df, sheet_name, startrow=None,
                           truncate_sheet=False,
                           **to_excel_kwargs):
        """
        Append a DataFrame [df] to existing Excel file [filename]
        into [sheet_name] Sheet.
        If [filename] doesn't exist, then this function will create it.

        Parameters:
          filename : File path or existing ExcelWriter
                     (Example: '/path/to/file.xlsx')
          df : dataframe to save to workbook
          sheet_name : Name of sheet which will contain DataFrame.
                       (default: 'Sheet1')
          startrow : upper left cell row to dump data frame.
                     Per default (startrow=None) calculate the last row
                     in the existing DF and write to the next row...
          truncate_sheet : truncate (remove and recreate) [sheet_name]
                           before writing DataFrame to Excel file
          to_excel_kwargs : arguments which will be passed to `DataFrame.to_excel()`
                            [can be dictionary]

        Returns: None

        (c) [MaxU](https://stackoverflow.com/users/5741205/maxu?tab=profile)
        """
        # ignore [engine] parameter if it was passed
        if 'engine' in to_excel_kwargs:
            to_excel_kwargs.pop('engine')

        # Python 2.x: define [FileNotFoundError] exception if it doesn't exist
        try:
            FileNotFoundError
        except NameError:
            FileNotFoundError = IOError

        try:
            writer = pd.ExcelWriter(filename, engine='openpyxl', mode='a')
            # try to open an existing workbook
            writer.book = load_workbook(filename)

            # get the last row in the existing Excel sheet
            # if it was not specified explicitly
            if startrow is None and sheet_name in writer.book.sheetnames:
                startrow = writer.book[sheet_name].max_row

            # truncate sheet
            if truncate_sheet and sheet_name in writer.book.sheetnames:
                # index of [sheet_name] sheet
                idx = writer.book.sheetnames.index(sheet_name)
                # remove [sheet_name]
                writer.book.remove(writer.book.worksheets[idx])
                # create an empty sheet [sheet_name] using old index
                writer.book.create_sheet(sheet_name, idx)

            # copy existing sheets
            writer.sheets = {ws.title: ws for ws in writer.book.worksheets}
        except FileNotFoundError:
            # file does not exist yet, we will create it
            writer = pd.ExcelWriter(filename, engine='openpyxl', mode = 'a')

        if startrow is None:
            startrow = 0

        # write out the new sheet
        df.to_excel(writer, sheet_name, startrow=startrow, **to_excel_kwargs)

        # save the workbook
        writer.save()

    append_df_to_excel('DF_Amazon Competitor Status_THANKYOU FARMER.xlsx', df_1, "Rice Toner Gift Set", index = False, header = False)

## Record Date ##
current_date = datetime.now()
# current_date = "%s-%s-%s-%s" % (current_time.month, current_time.day, current_time.year, current_time.time)
print("Current Date: ", current_date.strftime("%Y-%m-%d %H:%M:%S"))

## Opening page by search terms ##
print("Please enter the search terms you're looking for")
search_terms = input("")
base_URL = "https://www.amazon.com/s?k="
search_URL = base_URL + search_terms

driver.get(search_URL)
Request_page(search_URL)
html = driver.page_source
print(html)

## Adjust price range of competitors ##
print("Please select price range by number (Please enter the right matched texts exactly!!)")
print("1. Under $25 / 2. $25 to $50 / 3. $50 to $100 / 4. $100 to $200 / 5. $200 & Above")
price_range = input("")

if price_range == "1":
    try:
        # price_range_xpath = driver.find_element_by_xpath("//*[@id='p_36/1253950011']/div/span")
        # print(price_range_xpath.text)
        # price_range_xpath.click()

        price_range_url_xpath = driver.find_element_by_xpath("//*[@id='p_36/1253950011']")
        price_range_url = price_range_url_xpath.get_attribute("href")

        Request_page(price_range_url)

        renew_html = driver.page_source
        print(renew_html)

    # 구버전 UI
    except:
        # price_range_xpath = driver.find_element_by_xpath("//*[@id='p_36/1253503011']/div/span")
        # print(price_range_xpath.text)
        # price_range_xpath.click()

        price_range_url_xpath = driver.find_element_by_xpath("//*[@id='p_36/1253951011']/span/a")
        price_range_url = price_range_url_xpath.get_attribute("href")

        Request_page(price_range_url)

        renew_html = driver.page_source
        print(renew_html)

elif price_range == "2":
    try:
        # price_range_xpath = driver.find_element_by_xpath("//*[@id='p_36/1253951011']/div/span")
        # print(price_range_xpath.text)
        # price_range_xpath.click()

        price_range_url_xpath = driver.find_element_by_xpath("//*[@id='p_36/1253951011']/span/a")
        price_range_url = price_range_url_xpath.get_attribute("href")

        Request_page(price_range_url)

        renew_html = driver.page_source
        print(renew_html)

    # 구버전 UI
    except:
        # price_range_xpath = driver.find_element_by_xpath("//*[@id='p_36/1253504011']/span/a/span")
        # print(price_range_xpath.text)
        # price_range_xpath.click()

        price_range_url_xpath = driver.find_element_by_css_selector("#p_36\/1253951011 > span > a")
        price_range_url = price_range_url_xpath.get_attribute("href")

        Request_page(price_range_url)

        renew_html = driver.page_source
        print(renew_html)

elif price_range == "3":
    try:
        # price_range_xpath = driver.find_element_by_xpath("//*[@id='p_36/1253952011']/div/span")
        # print(price_range_xpath.text)
        # price_range_xpath.click()

        price_range_url_xpath = driver.find_element_by_xpath("//*[@id='p_36/1253952011']")
        price_range_url = price_range_url_xpath.get_attribute("href")

        Request_page(price_range_url)

        renew_html = driver.page_source
        print(renew_html)

    # 구버전 UI
    except:
        # price_range_xpath = driver.find_element_by_xpath("//*[@id='p_36/1253505011']/span/a/span")
        # print(price_range_xpath.text)
        # price_range_xpath.click()

        price_range_url_xpath = driver.find_element_by_xpath("//*[@id='p_36/1253952011']")
        price_range_url = price_range_url_xpath.get_attribute("href")

        Request_page(price_range_url)

        renew_html = driver.page_source
        print(renew_html)

elif price_range == "4":
    try:
        # price_range_xpath = driver.find_element_by_xpath("//*[@id='p_36/1253953011']/div/span")
        # print(price_range_xpath.text)
        # price_range_xpath.click()

        price_range_url_xpath = driver.find_element_by_xpath("/// *[ @ id = 'p_36/1253953011']")
        price_range_url = price_range_url_xpath.get_attribute("href")

        Request_page(price_range_url)

        renew_html = driver.page_source
        print(renew_html)

    # 구버전 UI
    except:
        # price_range_xpath = driver.find_element_by_xpath("//*[@id='p_36/1253506011']/span/a/span")
        # print(price_range_xpath.text)
        # price_range_xpath.click()

        price_range_url_xpath = driver.find_element_by_xpath("/// *[ @ id = 'p_36/1253953011']")
        price_range_url = price_range_url_xpath.get_attribute("href")

        Request_page(price_range_url)

        renew_html = driver.page_source
        print(renew_html)



# elif price_range == "5":
#     try:
#         price_range_xpath = driver.find_element_by_xpath("//*[@id='p_36/1253954011']/div/span")
#         print(price_range_xpath.text)
#         price_range_xpath.click()
#         html = "https://www.amazon.com/s?k=SSD&i=electronics&rh=n%3A172282%2Cp_36%3A1253507011&dc&qid=1640236403&rnid=386442011&ref=sr_nr_p_36_5"
#         soup = Request_page(html)
#         renew_html = driver.page_source
#         print(renew_html)
#     # 구버전 UI
#     except:
#         price_range_xpath = driver.find_element_by_xpath("//*[@id='p_36/1253507011']/span/a/span")
#         print(price_range_xpath.text)
#         price_range_xpath.click()
#         html = "https://www.amazon.com/s?k=SSD&i=electronics&rh=n%3A172282%2Cp_36%3A1253507011&dc&qid=1640236403&rnid=386442011&ref=sr_nr_p_36_5"
#         soup = Request_page(html)
#         renew_html = driver.page_source
#         print(renew_html)

# //*[@id="p_36/1253503011"]/span/a/span - Under $25
# //*[@id="p_36/1253504011"]/span/a/span - $25 to $50
# //*[@id="p_36/1253505011"]/span/a/span - $50 to $100
# //*[@id="p_36/1253506011"]/span/a/span - $100 to $200
# //*[@id="p_36/1253507011"]/span/a/span - $200 & Above

## Collecting URL of competitors ##
# sp_button_xpath = []
# sp_button = []
# url_collector_xpath = []
# url_collector_href = []
# sp_url_collector_xpath = []
# sp_url_collector_href = []
# asin = []
#
# # total title count: 60ea
# # for j in range(61):
# time.sleep(2)
#
# ## Extract only title and asin info ##
# title_list = []
#
# soup = BeautifulSoup(html, 'html.parser')
# divs = soup.find_all('div')
# # spans = soup.find_all('span',"a-size-base-plus a-color-base a-text-normal")
#
# asin_list = [div['data-asin'] for div in divs if div.has_attr('data-asin')]
#
# # title_list = [span['a-size-base-plus a-color-base a-text-normal'] for span in spans if span.has_attr('a-size-base-plus a-color-base a-text-normal')]
# for a in soup.find_all('span',"a-size-base-plus a-color-base a-text-normal"):
#     title_list.append(a.text)
#
# asin_xpath = driver.find_element_by_xpath("//*[@id='search']/div[1]/div[1]/div/span[3]/div[2]/div["+str(j+2)+"]")

## Extract only title and asin info ##
asin_list = []
title_list = []

soup = BeautifulSoup(html, 'html.parser')
divs = soup.find_all('div')
# spans = soup.find_all('span',"a-size-base-plus a-color-base a-text-normal")

asin_list = [div['data-asin'] for div in divs if div.has_attr('data-asin')]

# title_list = [span['a-size-base-plus a-color-base a-text-normal'] for span in spans if span.has_attr('a-size-base-plus a-color-base a-text-normal')]
for a in soup.find_all('span',"a-size-base-plus a-color-base a-text-normal"):
    title_list.append(a.text)

# asin_xpath = driver.find_element_by_xpath("//*[@id='search']/div[1]/div[1]/div/span[3]/div[2]/div["+str(j+2)+"]")
print("ASIN: ", asin_list)
print("Title: ", title_list)

save_excel(asin_list, title_list)

# asin_info = asin_xpath.text
# print(asin_info)
#
# asin.append(asin_info[j])
# print(asin)

# sponsored 버튼 여부로 sp / organic 구분
#     try:
#         # SP 버튼 먼저 찾아서 구분
#         time.sleep(2)
#         sp_button_xpath = driver.find_element_by_xpath("//*[@id='search']/div[1]/div[1]/div/span[3]/div[2]/div["+str(j)+"]/div/div/div/div/div/div/div/div[2]/div/div/div[1]/div/span")
#         if sp_url_collector_xpath.text == "Sponsored":
#             sp_button.append("Sponsored Ad")
#             print("This is Sponsored Ad")
#
#             # SP xpath 넘버링중 끊기는 부분 처리
#             try:
#                 url_collector_xpath = driver.find_element_by_xpath("//*[@id='search']/div[1]/div[1]/div/span[3]/div[2]/div["+str(j)+"]/div/div/div/div/div/div/div/div[2]/div/div/div[1]/h2/a")
#                 for t in url_collector_xpath:
#                     url_collector_href.append(url_collector_xpath[t].get_attribute("href"))
#                     print("url_collector_xpath: ", url_collector_xpath)
#                     print("url_collector_href: ", url_collector_href)
#
#             except:
#                 pass
#
#     except:
#         time.sleep(2)
#         sp_button.append("Organic")
#         print("This is Organic")
#
#         try:
#             url_collector_xpath = driver.find_element_by_xpath("//*[@id='search']/div[1]/div[1]/div/span[3]/div[2]/div["+str(j)+"]/div/div/div/div/div/div[2]/div/div/div[1]/h2/a")
#             for t in url_collector_xpath:
#                 url_collector_href.append(url_collector_xpath[t].get_attribute("href"))
#                 print("url_collector_xpath: ", url_collector_xpath)
#                 print("url_collector_href: ", url_collector_href)
#
#         except:
#             pass

# Organic 타이틀 xpath (sp, organic 구분 x)
# //*[@id="search"]/div[1]/div[1]/div/span[3]/div[2]/div[4]/div/div/div/div/div/div[2]/div/div/div[1]/h2


# 타이틀 xpath 자체로 sp / organic 구분
#     try:
#         time.sleep(2)
#         url_collector_xpath = driver.find_element_by_xpath("//*[@id='search']/div[1]/div[1]/div/span[3]/div[2]/div["+str(j+4)+"]/div/div/div/div/div/div[2]/div/div/div[1]/h2/a")
#         print("url_collector_xpath: ", url_collector_xpath)
#
#     except:
#         try:
#             time.sleep(2)
#             sp_url_collector_xpath = driver.find_element_by_xpath("//*[@id='search']/div[1]/div[1]/div/span[3]/div[2]/div["+str(j+2)+"]/div/div/div/div/div/div/div[2]/div/div/div[1]/h2")
#             print("SP_url_collector_xpath: ", sp_url_collector_xpath)
#
#         except:
#             time.sleep(2)
#             sp_url_collector_xpath = driver.find_element_by_xpath("//*[@id='search']/div[1]/div[1]/div/span[3]/div[2]/div["+str(j+2)+"]/div/span/div/div/div/div/div[2]/div[2]/div/div/div[1]/h2/a")
#             print("SP_url_collector_xpath: ", sp_url_collector_xpath)
#
# # //*[@id="search"]/div[1]/div[1]/div/span[3]/div[2]/div[2]/div/div/div/div/div/div/div/div[2]/div/div/div[1]/h2
# # //*[@id="search"]/div[1]/div[1]/div/span[3]/div[2]/div[2]/div/span/div/div/div/div/div[2]/div[2]/div/div/div[1]/h2/a
#
# for k in url_collector_xpath:
#     url_collector_href.append(k.get_attribute("href"))
#     print("url_collector_href: ", url_collector_href)
#
# for y in sp_url_collector_xpath:
#     sp_url_collector_href.append(y.get_attribute("href"))
#     print("SP_url_collector_href: ", sp_url_collector_href)

# ## Main ASIN URL ##
# brand = []
# asin = []
# title = []
# price = []
# ratings = []
# rank = []
#
# for a in url_collector_href:
#
#     driver.get(a)
#     time.sleep(5)
#     Request_page(a)
#     collected_html = driver.page_source
#     print(collected_html)
#
#     sp_button = sp_button
#     brand = driver.find_element_by_xpath("//*[@id='productDetails_techSpec_section_2']/tbody/tr[1]/td").text
#     asin = driver.find_element_by_xpath("//*[@id='productDetails_techSpec_section_2']/tbody/tr[11]/td")
#     title = driver.find_element_by_xpath("//*[@id='productTitle']").text
#     price = driver.find_element_by_xpath("//*[@id='corePrice_desktop']/div/table/tbody/tr/td[2]/span[1]").text
#     rating = driver.find_element_by_xpath("//*[@id='acrCustomerReviewText']").text
#     rank = driver.find_element_by_xpath("//*[@id='productDetails_detailBullets_sections1']/tbody/tr[2]/td/span/span").text

    # print("SP Button", sp_button)
    # print("Brand: ", brand)
    # print("ASIN", asin_list)
    # print("Title: ", title_list)
    # print("Price: ", price)
    # print("Ratings: ", rating)
    # print("Rank: ", rank)

    # save_excel(sp_button, brand, asin, title, price, rating, rank)
# save_excel(asin_list, title_list)

# # UTC time NOW # 한국시간에서 -9
# utc_now = datetime.utcnow()
# print(utc_now)

# scheduller ##
# # schedule.every(3).seconds.do(playing)
# schedule.every(1).day.at("20:13").do(playing)
#
# while True:
#     schedule.run_pending()
#     time.sleep(1)
#
# # ## Setting time zone1 ##
# import arrow
# from datetime import datetime
#
# now = datetime.now()
# atime = arrow.get(now)
# print(now)
# print (atime)
#
# eastern = atime.to('US/Eastern')
# print (eastern)
# print (eastern.datetime)

# data-asin 데이터만 extract하기
# https: // stackoverflow.com / questions / 56964475 / how - can - i - extract - all - div - that - have - a - specificl - element - in -it - that - its -not -cla / 56965221
